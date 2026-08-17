"""One-off: filled translation xlsx -> translations/manual_quotes.json.

    .venv/Scripts/python scripts/import_translations.py <filled.xlsx>

Reads the "Citatos" sheet of the export produced 2026-08-17 (columns located by
header name, so reordering is fine), keeps rows where "Vertimas LV" or
"Vertimas ET" was filled, and writes the committed translations file that
rossi_reviews/manual_translations.py loads at build time. Needs openpyxl
(requirements-dev.txt). Safe to re-run; the output is fully regenerated.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from rossi_reviews.transform import truncate_text

OUT = Path(__file__).resolve().parents[1] / "translations" / "manual_quotes.json"
MAX_LEN = 200  # FEATURED_MAX_LEN — keep translations within the same cap as quotes

COLUMNS = {
    "product_id": "product_id",
    "Autorius": "author",
    "Įvert.": "rating",
    "Citata LT": "source_text",
    "Vertimas LV": "lv",
    "Vertimas ET": "et",
}


def main(xlsx_path: str) -> int:
    ws = load_workbook(xlsx_path, read_only=True)["Citatos"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {}
    for name, key in COLUMNS.items():
        if name not in header:
            raise SystemExit(f"column {name!r} not found in the Citatos sheet")
        idx[key] = header.index(name)

    products: dict[str, dict] = {}
    skipped_same = 0
    for row in rows:
        pid = str(row[idx["product_id"]] or "").strip()
        source = str(row[idx["source_text"]] or "").strip()
        if not pid or not source:
            continue
        entry: dict[str, object] = {}
        for lang in ("lv", "et"):
            text = str(row[idx[lang]] or "").strip()
            if not text:
                continue
            if text == source:  # untouched copy of the LT text, not a translation
                skipped_same += 1
                continue
            entry[lang] = truncate_text(text, MAX_LEN)
        if not entry:
            continue
        rating = row[idx["rating"]]
        products[pid] = {
            "source_text": source,
            "author": str(row[idx["author"]] or "").strip() or None,
            "rating": int(rating) if rating is not None else None,
            **entry,
        }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    doc = {
        "generated_at": datetime.now(ZoneInfo("Europe/Vilnius")).isoformat(timespec="seconds"),
        "products": dict(sorted(products.items())),
    }
    OUT.write_text(
        json.dumps(doc, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    lv = sum(1 for p in products.values() if "lv" in p)
    et = sum(1 for p in products.values() if "et" in p)
    print(f"wrote {OUT}: {len(products)} products ({lv} LV, {et} ET)")
    if skipped_same:
        print(f"warning: {skipped_same} cells identical to the LT text — skipped (not translations)")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    raise SystemExit(main(sys.argv[1]))
